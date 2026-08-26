# Recalculate operations-v0.1.0.xlsx with Excel and report formula health.
import os
import shutil
import tempfile
import time

from openpyxl import load_workbook
from openpyxl.formula import Tokenizer
from openpyxl.utils.cell import range_boundaries

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, "docs", "operations-v0.1.0.xlsx")
ERRORS = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NULL!", "#NUM!")


def recalculate():
    import win32com.client

    tmp = tempfile.mkdtemp(prefix="pet-xlsx-")
    src = os.path.join(tmp, "in.xlsx")
    shutil.copy2(XLSX, src)
    abs_src = os.path.abspath(src)
    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.AskToUpdateLinks = False
    try:
        wb = excel.Workbooks.Open(abs_src, UpdateLinks=0)
        excel.Calculation = -4105  # xlCalculationAutomatic
        excel.CalculateFullRebuild()
        wb.Save()
        wb.Close(False)
    finally:
        excel.Quit()
        time.sleep(0.4)
    shutil.copy2(src, XLSX)
    shutil.rmtree(tmp, ignore_errors=True)


def inspect():
    formulas = load_workbook(XLSX, data_only=False)
    values = load_workbook(XLSX, data_only=True)
    formula_count = 0
    missing_cache = 0
    error_count = 0
    error_cells = []
    external = 0
    formula_cells = {}
    for idx, sheet in enumerate(formulas.worksheets):
        value_sheet = values.worksheets[idx]
        for row in sheet.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str) or not cell.value.startswith("="):
                    continue
                formula_count += 1
                formula_cells[(sheet.title, cell.coordinate)] = cell.value
                if "!" in cell.value and "[" in cell.value:
                    external += 1
                cached = value_sheet[cell.coordinate].value
                if cached is None:
                    missing_cache += 1
                    continue
                text = str(cached)
                if any(err in text for err in ERRORS):
                    error_count += 1
                    error_cells.append(f"{sheet.title}!{cell.coordinate}={text}")
    print(f"formulas={formula_count}")
    print(f"missing_cache={missing_cache}")
    print(f"formula_errors={error_count}")
    print(f"external_refs={external}")
    graph = {key: set() for key in formula_cells}
    for key, formula in formula_cells.items():
        current_sheet, _ = key
        for token in Tokenizer(formula).items:
            if token.subtype != "RANGE":
                continue
            raw = token.value
            target_sheet, address = current_sheet, raw
            if "!" in raw:
                sheet_part, address = raw.rsplit("!", 1)
                target_sheet = sheet_part.strip("'").replace("''", "'")
            if "[" in address or "," in address:
                continue
            try:
                min_col, min_row, max_col, max_row = range_boundaries(address.replace("$", ""))
            except ValueError:
                continue
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    coordinate = formulas.worksheets[0].cell(row=row, column=col).coordinate
                    dependency = (target_sheet, coordinate)
                    if dependency in formula_cells:
                        graph[key].add(dependency)
    visiting, visited = set(), set()

    def visit(node):
        if node in visiting:
            return True
        if node in visited:
            return False
        visiting.add(node)
        for dependency in graph[node]:
            if visit(dependency):
                return True
        visiting.remove(node)
        visited.add(node)
        return False

    circular = any(visit(node) for node in graph if node not in visited)
    print(f"circular_refs={1 if circular else 0}")
    if error_cells:
        print("errors:")
        print("\n".join(error_cells[:50]))
    if formula_count <= 0:
        raise SystemExit("no formulas found")
    if missing_cache > 0:
        raise SystemExit("formula caches are missing")
    if error_count > 0:
        raise SystemExit("formula errors present")
    if external > 0:
        raise SystemExit("external workbook refs present")
    if circular:
        raise SystemExit("circular formula references present")
    print("ok")


if __name__ == "__main__":
    print("recalculating with Excel COM")
    recalculate()
    inspect()
