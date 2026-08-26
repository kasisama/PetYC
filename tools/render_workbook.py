# Render every used cell of the recalculated workbook, paging long sheets.
import os
import sys

from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, "docs", "operations-v0.1.0.xlsx")
OUT_DIR = os.path.join(ROOT, "docs", "simulations", "workbook-sheets")
ROWS_PER_PAGE = 40
MAX_COLS = 14
ERRORS = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NULL!", "#NUM!")


def font(size, bold=False):
    for path in (r"C:\Windows\Fonts\msyhbd.ttc" if bold else r"C:\Windows\Fonts\msyh.ttc", r"C:\Windows\Fonts\simhei.ttf"):
        if os.path.exists(path):
            try:
                return ImageFont.truetype(path, size)
            except OSError:
                continue
    return ImageFont.load_default()


def cell_text(value):
    if value is None:
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return f"{value:.2f}"
    return str(value).replace("\n", " ")


def render_page(title, header, rows, path):
    header_font = font(16, True)
    body_font = font(13)
    all_rows = [header] + rows
    col_count = max(len(header), max((len(r) for r in rows), default=1))
    widths = [90] * col_count
    for row in all_rows:
        for i, value in enumerate(row):
            text = cell_text(value)[:48]
            visual = sum(16 if ord(ch) > 127 else 8 for ch in text)
            widths[i] = max(widths[i], min(360, visual + 20))
    width = min(2600, sum(widths) + 40)
    row_h = 28
    height = 48 + row_h * len(all_rows) + 16
    img = Image.new("RGB", (width, height), "#F7F5F0")
    draw = ImageDraw.Draw(img)
    draw.rectangle((0, 0, width, 40), fill="#1F4E79")
    draw.text((16, 8), title, font=header_font, fill="white")
    y = 44
    for ridx, row in enumerate(all_rows):
        x = 8
        bg = "#1F4E79" if ridx == 0 else ("#FFFFFF" if ridx % 2 else "#EEF3F8")
        fg = "white" if ridx == 0 else "#1A1A1A"
        draw.rectangle((8, y, width - 8, y + row_h - 1), fill=bg)
        for i in range(col_count):
            text = cell_text(row[i] if i < len(row) else "")[:48]
            if text.startswith("="):
                raise SystemExit(f"formula leaked into render: {title} {text}")
            if any(err in text for err in ERRORS):
                raise SystemExit(f"error value in render: {title} {text}")
            draw.text((x + 8, y + 6), text, font=body_font, fill=fg)
            x += widths[i]
        y += row_h
    img.save(path)
    return os.path.getsize(path)


def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    for name in os.listdir(OUT_DIR):
        if name.endswith(".png"):
            os.remove(os.path.join(OUT_DIR, name))
    wb = load_workbook(XLSX, data_only=True)
    if len(wb.sheetnames) != 20:
        print("expected 20 sheets, got", len(wb.sheetnames), file=sys.stderr)
        return 1
    rendered = []
    for idx, sheet in enumerate(wb.worksheets, 1):
        rows = []
        max_col = min(sheet.max_column or 1, MAX_COLS)
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, max_col=max_col, values_only=True):
            rows.append(list(row))
        if not rows:
            print("empty sheet", sheet.title, file=sys.stderr)
            return 1
        header, body = rows[0], rows[1:]
        pages = max(1, (len(body) + ROWS_PER_PAGE - 1) // ROWS_PER_PAGE)
        safe = "".join(ch if ch.isalnum() or ch in "-_ " else "_" for ch in sheet.title)
        for page in range(pages):
            chunk = body[page * ROWS_PER_PAGE : (page + 1) * ROWS_PER_PAGE]
            suffix = f"-p{page+1:02d}" if pages > 1 else ""
            path = os.path.join(OUT_DIR, f"{idx:02d}-{safe}{suffix}.png")
            size = render_page(f"{sheet.title}{suffix}", header, chunk, path)
            print("rendered", path, "bytes", size, "rows", 1 + len(chunk), "/", len(rows))
            rendered.append(path)
            if size < 3000:
                print("tiny render", path, file=sys.stderr)
                return 1
    print("ok", len(rendered), "images from", len(wb.sheetnames), "sheets")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
