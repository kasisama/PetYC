# Generate the 20-sheet operations workbook from config_v0.1.0.json.
import json
import os
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CFG = os.path.join(ROOT, "config", "defaults", "config_v0.1.0.json")
SIM = os.path.join(ROOT, "docs", "simulations", "season70-10000.json")
OUT = os.path.join(ROOT, "docs", "operations-v0.1.0.xlsx")

data = json.load(open(CFG, encoding="utf-8"))
sim = {}
if os.path.exists(SIM):
    sim = json.load(open(SIM, encoding="utf-8"))

pets = data.get("pet_species", [])
items = data.get("items", [])
currencies = data.get("currencies", [])
shops = data.get("shop_items", [])
adv_shops = data.get("adventure_shop_items", [])
maps_ = data.get("adventure_maps", [])
zones = data.get("adventure_zones", [])
monsters = data.get("adventure_monsters", [])
skills = data.get("adventure_skills", [])
equips = data.get("equipment_templates", [])
affixes = data.get("equipment_affixes", [])
recipes = data.get("equipment_recipes", [])
materials = data.get("equipment_recipe_materials", [])
loots = data.get("adventure_loot_entries", [])
events = data.get("live_events", [])
tracks = data.get("reward_tracks", [])
levels = data.get("adventure_levels", [])
evo_rules = data.get("pet_evolution_rules", [])
evo_costs = data.get("pet_evolution_costs", [])
images = data.get("images", [])

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Microsoft YaHei")
BODY_FONT = Font(name="Microsoft YaHei", size=11)
OK_FILL = PatternFill("solid", fgColor="C6EFCE")
BAD_FILL = PatternFill("solid", fgColor="FFC7CE")
WARN_FILL = PatternFill("solid", fgColor="FFEB9C")
THIN = Border(
    left=Side(style="thin", color="D0D7DE"),
    right=Side(style="thin", color="D0D7DE"),
    top=Side(style="thin", color="D0D7DE"),
    bottom=Side(style="thin", color="D0D7DE"),
)


def g(obj, *keys, default=""):
    if not isinstance(obj, dict):
        return default
    for key in keys:
        if key in obj and obj[key] not in (None, ""):
            return obj[key]
    return default


def n(obj, *keys, default=0):
    value = g(obj, *keys, default=default)
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


wb = Workbook()


def style_header(ws, cols):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(cols)}{ws.max_row}"
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=cols):
        for cell in row:
            cell.font = BODY_FONT
            cell.border = THIN
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    def display_len(text):
        return sum(2 if ord(ch) > 127 else 1 for ch in str(text or ""))

    for col in range(1, cols + 1):
        letter = get_column_letter(col)
        width = 14
        for cell in ws[letter]:
            width = max(width, min(48, display_len(cell.value) + 4))
        ws.column_dimensions[letter].width = width
    ws.row_dimensions[1].height = 22
    ws.sheet_properties.tabColor = "1F4E79"


def add_sheet(name, headers, rows, number_cols=None, formulas=None):
    ws = wb.create_sheet(name[:31])
    ws.append(headers)
    number_cols = set(number_cols or [])
    for ridx, row in enumerate(rows, start=2):
        values = list(row)
        if formulas and ridx in formulas:
            for col, formula in formulas[ridx].items():
                values[col] = formula
        ws.append(values)
        for col in number_cols:
            cell = ws.cell(ridx, col + 1)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.00" if isinstance(cell.value, float) else "0"
    if not rows:
        ws.append(["（无数据）"] + [""] * (len(headers) - 1))
    style_header(ws, len(headers))
    return ws


# 1 总览
overview = wb.active
overview.title = "总览"
overview.append(["指标", "数值", "公式/来源", "目标"])
for row in [
    ("宠物形态", len(pets), "COUNTA(宠物形态!A:A)-1", "20"),
    ("谱系", len({g(p, "family_key", "FamilyKey") for p in pets}), "去重", "5"),
    ("统一物品", len(items), "COUNTA(统一物品!A:A)-1", "45"),
    ("货币", len(currencies), "COUNTA(货币与经济!A:A)-1", "3"),
    ("地图", len(maps_), "COUNTA(地图与区域!A:A)-1", "3"),
    ("区域", len(zones), "COUNTA", "12"),
    ("怪物", len(monsters), "COUNTA(怪物!A:A)-1", ">=33"),
    ("技能", len(skills), "COUNTA(技能!A:A)-1", ">=30"),
    ("装备", len(equips), "COUNTA(装备模板!A:A)-1", ">=30"),
    ("词条", len(affixes), "COUNTA(词条池!A:A)-1", ">=30"),
    ("配方", len(recipes), "COUNTA(配方与分解!A:A)-1", ">=12"),
    ("奖励轨", len(tracks), "COUNTA('20档奖励轨'!A:A)-1", "20"),
    ("标准日收入下限", 250, "输入", "250"),
    ("标准日收入上限", 350, "输入", "350"),
    ("标准日消耗下限", 180, "输入", "180"),
    ("标准日消耗上限", 280, "输入", "280"),
    ("日净流入中值", "=(B14+B15)/2-(B16+B17)/2", "(收入中值-消耗中值)", "约 70"),
]:
    overview.append(list(row))
style_header(overview, 4)
overview.conditional_formatting.add("B2:B13", CellIsRule(operator="lessThan", formula=["1"], fill=BAD_FILL))

# 2 宠物形态
pet_rows = []
for i, p in enumerate(pets, start=2):
    pet_rows.append([
        g(p, "key", "Key"),
        g(p, "name", "Name"),
        g(p, "family_key", "FamilyKey"),
        g(p, "stage", "Stage"),
        n(p, "wisdom", "Wisdom"),
        n(p, "strength", "Strength"),
        n(p, "defense", "Defense"),
        f"=E{i}+F{i}+G{i}",
        f"=IF(H$2=0,0,H{i}-H$2)",
        g(p, "adoptable", "Adoptable"),
        g(p, "image", "Image") or "待绘制",
    ])
add_sheet("宠物形态", ["key", "name", "family", "stage", "wisdom", "strength", "defense", "budget", "delta", "adoptable", "image"], pet_rows, number_cols={4, 5, 6})

# 3 进化链与消耗
cost_map = {}
for row in evo_costs:
    key = g(row, "evolution_key", "EvolutionKey")
    cost_map.setdefault(key, []).append(f"{g(row, 'item_key', 'ItemKey')}*{int(n(row, 'quantity', 'Quantity'))}")
evo_rows = []
for i, rule in enumerate(evo_rules, start=2):
    key = g(rule, "key", "Key")
    evo_rows.append([
        key,
        g(rule, "from_form_key", "FromFormKey"),
        g(rule, "to_form_key", "ToFormKey"),
        g(rule, "branch_label", "BranchLabel"),
        n(rule, "required_growth", "RequiredGrowth"),
        n(rule, "required_affection", "RequiredAffection"),
        " / ".join(cost_map.get(key, [])),
        g(rule, "enabled", "Enabled"),
        f'=IF(E{i}<0,"异常","ok")',
    ])
add_sheet("进化链与消耗", ["rule", "from", "to", "branch", "growth", "affection", "items", "enabled", "check"], evo_rows, number_cols={4, 5})

# 4 等级曲线
level_rows = []
for i, lv in enumerate(levels, start=2):
    level_rows.append([
        int(n(lv, "level", "Level", default=i - 1)),
        int(n(lv, "xp_to_next", "XPToNext")),
        "=B2" if i == 2 else f"=C{i-1}+B{i}",
        f"=IF(C{i}=0,1,C{i}/220)",
    ])
ws_level = add_sheet("等级曲线", ["level", "xp_to_next", "累计经验", "预计达到日"], level_rows, number_cols={0, 1})
ws_level.conditional_formatting.add("D2:D26", CellIsRule(operator="greaterThan", formula=["80"], fill=WARN_FILL))

# 5 统一物品
item_rows = [[g(it, "key", "Key"), g(it, "name", "Name"), g(it, "category", "Category"), g(it, "rarity", "Rarity"), g(it, "usage", "Usage"), n(it, "sell_price", "SellPrice"), g(it, "image", "Image") or "待绘制"] for it in items]
ws_items = add_sheet("统一物品", ["key", "name", "category", "rarity", "usage", "sell_price", "image"], item_rows, number_cols={5})
ws_items.conditional_formatting.add('A2:A200', FormulaRule(formula=['A2="season_token"'], fill=BAD_FILL))

# 6 货币与经济
cur_rows = []
for i, cur in enumerate(currencies, start=2):
    key = g(cur, "key", "Key")
    source = {"primary_coin": "签到/打工", "journey_badge": "区域掉落", "season_token": "奖励轨"}.get(key, "")
    sink = {"primary_coin": "常规商店", "journey_badge": "调查商店", "season_token": "赛季商店"}.get(key, "")
    cur_rows.append([key, g(cur, "name", "Name"), g(cur, "description", "Description"), source, sink, g(cur, "enabled", "Enabled"), f'=IF(OR(D{i}="",E{i}=""),"缺口","闭环")'])
add_sheet("货币与经济", ["key", "name", "description", "source", "sink", "enabled", "loop"], cur_rows)

# 7 常规商店
normal_rows = []
for i, shop in enumerate(shops, start=2):
    normal_rows.append([g(shop, "name", "Name"), n(shop, "price", "Price"), g(shop, "shop_type", "ShopType"), n(shop, "stock", "Stock"), "primary_coin", f"=IF(B{i}<=0,\"异常\",\"ok\")"])
add_sheet("常规商店", ["name", "price", "shop_type", "stock", "currency", "check"], normal_rows, number_cols={1, 3})

# 8 赛季商店 + 调查商店都放，筛 season
season_rows = []
journey_rows_start = 2
for listing in adv_shops:
    season_rows.append([
        g(listing, "key", "Key"),
        g(listing, "name", "Name"),
        g(listing, "product_key", "ProductKey"),
        g(listing, "product_type", "ProductType"),
        n(listing, "price", "Price"),
        g(listing, "currency_key", "CurrencyKey"),
        g(listing, "limit_type", "LimitType"),
        n(listing, "limit_quantity", "LimitQuantity"),
        f'=IF(F{len(season_rows)+2}="","缺货币","ok")',
    ])
add_sheet("赛季商店", ["key", "name", "product", "type", "price", "currency", "limit_type", "limit_qty", "check"], season_rows, number_cols={4, 7})

# 9 地图与区域
map_lookup = {g(m, "key", "Key"): g(m, "name", "Name") for m in maps_}
zone_rows = []
for z in zones:
    zone_rows.append([g(z, "map_key", "MapKey"), map_lookup.get(g(z, "map_key", "MapKey"), ""), g(z, "key", "Key"), g(z, "name", "Name"), n(z, "recommended_level", "RecommendedLevel"), n(z, "hunger_cost", "HungerCost")])
add_sheet("地图与区域", ["map_key", "map_name", "zone_key", "zone_name", "level", "hunger"], zone_rows, number_cols={4, 5})

# 10 怪物
mon_rows = [[g(m, "key", "Key"), g(m, "name", "Name"), n(m, "level", "Level"), n(m, "max_health", "MaxHealth"), n(m, "attack", "Attack"), n(m, "defense", "Defense"), bool(g(m, "elite", "Elite", default=False))] for m in monsters]
add_sheet("怪物", ["key", "name", "level", "hp", "atk", "def", "elite"], mon_rows, number_cols={2, 3, 4, 5})

# 11 技能
skill_rows = [[g(s, "key", "Key"), g(s, "name", "Name"), n(s, "power_permille", "PowerPermille"), n(s, "accuracy_permille", "AccuracyPermille"), g(s, "effect_type", "EffectType")] for s in skills]
add_sheet("技能", ["key", "name", "power", "accuracy", "effect"], skill_rows, number_cols={2, 3})

# 12 装备模板
equip_rows = []
for i, e in enumerate(equips, start=2):
    equip_rows.append([
        g(e, "key", "Key"), g(e, "name", "Name"), g(e, "slot", "Slot"), g(e, "rarity", "Rarity"),
        n(e, "required_level", "RequiredLevel"), g(e, "salvage_item", "SalvageItem"), n(e, "salvage_quantity", "SalvageQuantity"),
        g(e, "affix_pool_key", "AffixPoolKey"), f"=G{i}",
    ])
add_sheet("装备模板", ["key", "name", "slot", "rarity", "level", "salvage_item", "salvage_qty", "affix_pool", "salvage_qty_formula"], equip_rows, number_cols={4, 6})

# 13 词条池
affix_rows = [[g(a, "key", "Key"), g(a, "pool_key", "PoolKey"), g(a, "attribute", "Attribute"), n(a, "min_value", "MinValue"), n(a, "max_value", "MaxValue"), n(a, "weight", "Weight")] for a in affixes]
add_sheet("词条池", ["key", "pool", "attr", "min", "max", "weight"], affix_rows, number_cols={3, 4, 5})

# 14 配方与分解
mat_map = {}
for row in materials:
    key = g(row, "equipment_key", "EquipmentKey")
    mat_map.setdefault(key, []).append((g(row, "item_name", "ItemName"), n(row, "quantity", "Quantity")))
sell = {g(it, "key", "Key") or g(it, "name", "Name"): n(it, "sell_price", "SellPrice", default=2) for it in items}
recipe_rows = []
for i, rec in enumerate(recipes, start=2):
    key = g(rec, "equipment_key", "EquipmentKey")
    mats = mat_map.get(key, [])
    craft_terms = "+".join(str(int(qty) * max(sell.get(name, 2), 2)) for name, qty in mats) or "0"
    salvage_item = ""
    salvage_qty = 0
    for e in equips:
        if g(e, "key", "Key") == key:
            salvage_item = g(e, "salvage_item", "SalvageItem")
            salvage_qty = n(e, "salvage_quantity", "SalvageQuantity")
    recipe_rows.append([
        key, g(rec, "blueprint_fragment_item", "BlueprintFragmentItem"), n(rec, "blueprint_fragments", "BlueprintFragments"),
        n(rec, "currency_cost", "CurrencyCost"), " / ".join(f"{a}*{int(b)}" for a, b in mats),
        f"={craft_terms}", salvage_item, salvage_qty, f"=H{i}*2", f"=IF(AND(F{i}>0,I{i}>=F{i}),\"套利\",\"ok\")",
    ])
ws_rec = add_sheet("配方与分解", ["equipment", "blueprint", "fragments", "coin", "materials", "craft_value", "salvage_item", "salvage_qty", "salvage_value", "loop_check"], recipe_rows, number_cols={2, 3, 7})
ws_rec.conditional_formatting.add("J2:J200", FormulaRule(formula=['J2="套利"'], fill=BAD_FILL))
ws_rec.conditional_formatting.add("J2:J200", FormulaRule(formula=['J2="ok"'], fill=OK_FILL))

# 15 掉落池
loot_rows = [[g(l, "pool_key", "PoolKey"), g(l, "reward_type", "RewardType"), g(l, "reward_key", "RewardKey"), n(l, "min_quantity", "MinQuantity"), n(l, "max_quantity", "MaxQuantity"), n(l, "weight", "Weight"), bool(g(l, "guaranteed", "Guaranteed", default=False))] for l in loots]
add_sheet("掉落池", ["pool", "type", "key", "min", "max", "weight", "guaranteed"], loot_rows, number_cols={3, 4, 5})

# 16 70 天活动
event_rows = [[g(e, "key", "Key"), g(e, "name", "Name"), g(e, "region", "Region"), g(e, "starts_at", "StartsAt"), g(e, "ends_at", "EndsAt"), g(e, "active", "Active")] for e in events]
add_sheet("70天活动", ["key", "name", "region", "start", "end", "active"], event_rows)

# 17 20 档奖励轨
track_rows = []
for i, t in enumerate(tracks, start=2):
    track_rows.append([
        g(t, "event_key", "EventKey"), int(n(t, "milestone", "Milestone")), g(t, "reward_type", "RewardType"),
        g(t, "reward_key", "RewardKey"), g(t, "reward_name", "RewardName"), int(n(t, "quantity", "Quantity")),
        f'=IF(OR(C{i}="",D{i}=""),"缺字段","ok")',
    ])
add_sheet("20档奖励轨", ["event", "milestone", "type", "key", "name", "qty", "check"], track_rows, number_cols={1, 5})

# 18 模拟结果
cohorts = sim.get("cohorts", {})
sim_rows = [["cohort", "income_p50", "spend_p50", "crafts_p50", "evolve_p50", "awaken_p50", "boss_join", "boss_clear", "badge_p50", "season_p50"]]
for name in ("low", "standard", "high"):
    row = cohorts.get(name, {})
    def r(value):
        try:
            return round(float(value), 2)
        except (TypeError, ValueError):
            return 0
    sim_rows.append([
        name, r(row.get("daily_income_p50", 0)), r(row.get("daily_spend_p50", 0)), r(row.get("crafts_p50", 0)),
        r(row.get("evolve_day_p50", 0)), r(row.get("awaken_day_p50", 0)), r(row.get("boss_join_rate", 0)),
        r(row.get("boss_clear_rate", 0)), r(row.get("journey_badge_p50", 0)), r(row.get("season_token_p50", 0)),
    ])
ws_sim = wb.create_sheet("模拟结果")
for r in sim_rows:
    ws_sim.append(r)
ws_sim.append([])
ws_sim.append(["标准收入是否达标", f'=AND(B3>=250,B3<=350)'])
ws_sim.append(["标准消耗是否达标", f'=AND(C3>=180,C3<=280)'])
ws_sim.append(["进化日是否达标", f'=AND(E3>=7,E3<=10)'])
ws_sim.append(["觉醒日是否达标", f'=AND(F3>=35,F3<=49)'])
ws_sim.append(["谱系战力比", round(float(sim.get("power_spread_ratio", 0) or 0), 3)])
ws_sim.append(["无统治谱系", "=AND(B10<1.25,B10>0)"])
level15 = sim.get("level_arrival", {}).get("15", {})
zones = sim.get("zone_unlock", {})
ws_sim.append(["Lv.15到达人数", int(level15.get("players", 0) or 0), "P50日", int(level15.get("day_p50", 0) or 0)])
ws_sim.append(["首图完成P50", int(zones.get("zone_04", {}).get("unlock_p50", 0) or 0), "目标<=20", "=B13<=20"])
ws_sim.append(["第二图进入P50", int(zones.get("zone_05", {}).get("unlock_p50", 0) or 0), "目标<=40", "=B14<=40"])
ws_sim.append(["第三图进入P50", int(zones.get("zone_09", {}).get("unlock_p50", 0) or 0), "目标<=70", "=B15<=70"])
ws_sim.append([])
ws_sim.append(["目标键", "模拟器结果"])
for key in ("income_ok", "spend_ok", "evolve_ok", "awaken_ok", "equip_ok", "no_ruler", "progress_ok", "boss_ok"):
    ws_sim.append([key, bool(sim.get("targets", {}).get(key, False))])
style_header(ws_sim, 10)
for cell in ws_sim[17]:
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws_sim.freeze_panes = "A2"
ws_sim.auto_filter.ref = "A1:J4"
for column in range(1, 11):
    ws_sim.column_dimensions[get_column_letter(column)].width = 17
for cell in ("B6", "B7", "B8", "B9", "B11", "D13", "D14", "D15"):
    ws_sim.conditional_formatting.add(cell, FormulaRule(formula=[f"{cell}=TRUE"], fill=OK_FILL))
    ws_sim.conditional_formatting.add(cell, FormulaRule(formula=[f"{cell}=FALSE"], fill=BAD_FILL))
ws_sim.conditional_formatting.add("B18:B25", FormulaRule(formula=["B18=TRUE"], fill=OK_FILL))
ws_sim.conditional_formatting.add("B18:B25", FormulaRule(formula=["B18=FALSE"], fill=BAD_FILL))

# 19 资产需求
asset_rows = []
for p in pets:
    asset_rows.append(["pet", g(p, "key", "Key"), "image", g(p, "image", "Image") or "待绘制"])
for z in zones:
    asset_rows.append(["zone", g(z, "key", "Key"), "image", g(z, "image", "Image") or "待绘制"])
for e in equips:
    asset_rows.append(["equipment", g(e, "key", "Key"), "image", g(e, "image", "Image") or "待绘制"])
for img in images:
    asset_rows.append(["mapped", g(img, "name", "Name"), "path", g(img, "path", "Path") or "空"])
add_sheet("资产需求", ["kind", "key", "field", "status"], asset_rows)

# 20 校验结果
checks = [
    ("地图数量", len(maps_), 3, f"=B2=C2"),
    ("区域数量", len(zones), 12, f"=B3=C3"),
    ("宠物形态", len(pets), 20, f"=B4=C4"),
    ("奖励轨", len(tracks), 20, f"=B5=C5"),
    ("技能>=30", len(skills), 30, f"=B6>=C6"),
    ("装备>=30", len(equips), 30, f"=B7>=C7"),
    ("词条>=30", len(affixes), 30, f"=B8>=C8"),
    ("配方>=12", len(recipes), 12, f"=B9>=C9"),
    ("货币=3", len(currencies), 3, f"=B10=C10"),
    ("season_token物品数", sum(1 for it in items if g(it, "key", "Key") == "season_token"), 0, f"=B11=C11"),
    ("赛季商店", sum(1 for s in adv_shops if g(s, "currency_key", "CurrencyKey") == "season_token"), 1, f"=B12>=C12"),
    ("调查商店", sum(1 for s in adv_shops if g(s, "currency_key", "CurrencyKey") == "journey_badge"), 1, f"=B13>=C13"),
    ("徽章掉落", sum(1 for l in loots if g(l, "reward_type", "RewardType") == "currency" and g(l, "reward_key", "RewardKey") == "journey_badge"), 1, f"=B14>=C14"),
    ("蓝图碎片来源", sum(1 for l in loots if g(l, "reward_type", "RewardType") == "blueprint_fragment"), 12, f"=B15>=C15"),
    ("模拟目标通过", sum(1 for value in sim.get("targets", {}).values() if value is True), 8, f"=B16=C16"),
]
ws_chk = add_sheet("校验结果", ["检查项", "实际", "目标", "通过"], [[a, b, c, d] for a, b, c, d in checks], number_cols={1, 2})
ws_chk.conditional_formatting.add("D2:D20", FormulaRule(formula=["D2=TRUE"], fill=OK_FILL))
ws_chk.conditional_formatting.add("D2:D20", FormulaRule(formula=["D2=FALSE"], fill=BAD_FILL))

# remove default extra if any leftover
while "Sheet" in wb.sheetnames:
    del wb[wb.sheetnames[wb.sheetnames.index("Sheet")]]

os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)
print("wrote", OUT, "sheets", len(wb.sheetnames))
print("sheet names:", ", ".join(wb.sheetnames))
